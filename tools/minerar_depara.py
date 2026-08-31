#!/usr/bin/env python3
"""Minera a base De/Para inicial a partir do historico ja classificado pelo contador.

Fase 0 do plano de engenharia. Ferramenta offline: le os arquivos entregues pelo
cliente e produz uma base candidata para revisao humana. Nao faz parte do webapp.

Ver docs/adr/0003-constituicao-base-depara.md para o racional.

Uso:
    python3 tools/minerar_depara.py [--dir arquivos-clickip/clickIP] [--out docs]
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pdfplumber

# Junho e o unico mes cuja planilha traz a coluna "favorecido / beneficiario".
# Ele define o conjunto autoritativo de fornecedores; os outros meses so reforcam
# a confianca. Ver docs/02-analise-arquivos-cliente.md secao 5.
MES_COM_FAVORECIDO = "CLICK SCM 062026.xlsx"

COL_DATA, COL_DEBITO, COL_VALOR, COL_HISTORICO, COL_CENTRO, COL_FAVORECIDO = 1, 2, 4, 5, 6, 10

HISTORICO_GENERICO = "SISPAG FORNECEDORES"

# Nomes chegam truncados em 30 caracteres em parte das fontes. Comparar por um
# prefixo mais curto une "EQUATORIAL PARA DISTRIBUIDORA" e a versao completa.
TAMANHO_CHAVE_NOME = 29

# "GRUPO MULTI S.A" e "GRUPO MULTI SA" sao o mesmo fornecedor; sem colapsar o
# sufixo societario eles viram duas regras concorrentes.
SUFIXOS_SOCIETARIOS = re.compile(r"\b(S A|SA|S S|LTDA|ME|EPP|EIRELI|MEI|S C|CIA)\b")


def normalizar_nome(valor: str) -> str:
    """Maiusculas, sem acento, sem pontuacao, espacos colapsados."""
    sem_acento = "".join(
        c for c in unicodedata.normalize("NFKD", str(valor)) if not unicodedata.combining(c)
    )
    return re.sub(r"\s+", " ", re.sub(r"[^A-Za-z0-9 ]", " ", sem_acento)).strip().upper()


def chave_nome(valor: str) -> str:
    """Chave de agrupamento: nome sem sufixo societario, truncado."""
    sem_sufixo = SUFIXOS_SOCIETARIOS.sub(" ", normalizar_nome(valor))
    return re.sub(r"\s+", " ", sem_sufixo).strip()[:TAMANHO_CHAVE_NOME]


def normalizar_documento(valor: str) -> str:
    """CPF/CNPJ reduzido a digitos. Devolve '' se nao tiver 11 ou 14 digitos."""
    digitos = re.sub(r"\D", "", str(valor or ""))
    return digitos if len(digitos) in (11, 14) else ""


@dataclass
class Fornecedor:
    """Observacoes acumuladas de um fornecedor ao longo do historico."""

    chave: str
    nomes: set[str] = field(default_factory=set)
    documentos: set[str] = field(default_factory=set)
    # (conta_debito, centro_custo) -> {competencias onde apareceu}
    destinos: dict[tuple[str, str], set[str]] = field(default_factory=lambda: defaultdict(set))

    @property
    def nome_canonico(self) -> str:
        """O nome mais longo observado, que e o menos truncado."""
        return max(self.nomes, key=len)

    @property
    def documento(self) -> str:
        return sorted(self.documentos)[0] if len(self.documentos) == 1 else ""

    @property
    def competencias(self) -> set[str]:
        return {c for meses in self.destinos.values() for c in meses}

    @property
    def contas(self) -> set[str]:
        return {conta for conta, _ in self.destinos}

    @property
    def centro_dominante(self) -> str:
        """O centro de custo mais observado. Ver classificar() para o porque."""
        por_centro: dict[str, int] = defaultdict(int)
        for (_, centro), meses in self.destinos.items():
            por_centro[centro] += len(meses)
        return max(por_centro, key=lambda c: (por_centro[c], c == "0001"))

    @property
    def centros_alternativos(self) -> list[str]:
        return sorted({c for _, c in self.destinos} - {self.centro_dominante})

    def classificar(self) -> str:
        """Confianca da regra.

        Ambiguidade de CONTA e bloqueante: o contador usa contas diferentes de
        proposito (uma por cota de consorcio, consumo vs uso de poste) e nao ha
        como escolher automaticamente.

        Ambiguidade de CENTRO DE CUSTO nao e bloqueante. Medimos que o centro nao
        e funcao do fornecedor - TERACOM aparece em 5 centros com uma unica conta -
        e nao e derivavel do relatorio Itau (a coluna "referencia da empresa" vem
        vazia em 164 de 284 linhas casadas). Entao a regra fixa a conta, propoe o
        centro dominante e deixa o ajuste para a tela de validacao.
        """
        if len(self.contas) > 1:
            return "AMBIGUO_CONTA"
        return "ALTA" if len(self.competencias) >= 3 else "MEDIA"


def competencia_de(data: str) -> str:
    """'02/01/2026' -> '2026-01'."""
    partes = str(data).split("/")
    return f"{partes[2]}-{partes[1]}" if len(partes) == 3 else "?"


def ler_planilhas(pasta: Path) -> tuple[dict[str, Fornecedor], dict]:
    """Le os XLSX de saida Fortes e acumula observacoes por fornecedor."""
    fornecedores: dict[str, Fornecedor] = {}
    stats = {"arquivos": [], "duplicatas": [], "linhas": 0, "sispag": 0, "sem_fornecedor": 0}
    vistos: dict[bytes, str] = {}

    def obter(chave: str) -> Fornecedor:
        return fornecedores.setdefault(chave, Fornecedor(chave=chave))

    caminho_junho = pasta / MES_COM_FAVORECIDO
    if not caminho_junho.exists():
        sys.exit(f"erro: {MES_COM_FAVORECIDO} nao encontrado em {pasta}")

    # Passo 1: junho define o conjunto de fornecedores, sem inferencia.
    for linha in linhas_de(caminho_junho):
        favorecido = linha[COL_FAVORECIDO]
        if not favorecido:
            stats["sem_fornecedor"] += 1
            continue
        f = obter(chave_nome(favorecido))
        f.nomes.add(normalizar_nome(favorecido))
        f.destinos[(str(linha[COL_DEBITO]), str(linha[COL_CENTRO]))].add(
            competencia_de(linha[COL_DATA])
        )

    conhecidos = set(fornecedores)

    # Uma conta de debito que junho usou para um unico fornecedor identifica esse
    # fornecedor sem ambiguidade. Isso permite reforcar a confianca com os meses
    # que nao tem coluna de favorecido, sem inventar fornecedor novo.
    dono_da_conta: dict[str, set[str]] = defaultdict(set)
    for chave, f in fornecedores.items():
        for conta in f.contas:
            dono_da_conta[conta].add(chave)
    conta_exclusiva = {c: next(iter(d)) for c, d in dono_da_conta.items() if len(d) == 1}

    # Passo 2: os outros meses reforcam a confianca por duas vias.
    #  (a) sufixo do historico depois do ultimo " - ", aceito so quando casa com
    #      um fornecedor que junho ja provou existir - o sufixo tambem carrega
    #      complementos que nao sao nome ("Santarem", "Uso de Postes");
    #  (b) conta de debito exclusiva de um fornecedor.
    for arquivo in sorted(pasta.glob("CLICK SCM *.xlsx")):
        conteudo = arquivo.read_bytes()
        if conteudo in vistos:
            stats["duplicatas"].append((arquivo.name, vistos[conteudo]))
            continue
        vistos[conteudo] = arquivo.name
        contagem = 0
        for linha in linhas_de(arquivo):
            contagem += 1
            stats["linhas"] += 1
            historico = str(linha[COL_HISTORICO] or "").strip()
            if historico == HISTORICO_GENERICO:
                stats["sispag"] += 1
            if arquivo.name == MES_COM_FAVORECIDO:
                continue
            conta = str(linha[COL_DEBITO])
            chave = ""
            if " - " in historico:
                candidato = chave_nome(historico.rsplit(" - ", 1)[1])
                if candidato in conhecidos:
                    chave = candidato
            if not chave:
                chave = conta_exclusiva.get(conta, "")
            if chave:
                fornecedores[chave].destinos[(conta, str(linha[COL_CENTRO]))].add(
                    competencia_de(linha[COL_DATA])
                )
            else:
                stats["nao_atribuidas"] = stats.get("nao_atribuidas", 0) + 1
        stats["arquivos"].append((arquivo.name, contagem))

    return fornecedores, stats


def linhas_de(caminho: Path) -> list[tuple]:
    ws = openpyxl.load_workbook(caminho, data_only=True).active
    return [r for r in ws.iter_rows(min_row=2, values_only=True) if r[COL_DATA]]


def ler_documentos_itau(pasta: Path) -> dict[str, str]:
    """Extrai favorecido -> CPF/CNPJ dos relatorios Itau, para dar chave canonica."""
    documentos: dict[str, str] = {}
    for pdf_path in sorted(pasta.glob("Relatorio ITAU*.pdf")):
        with pdfplumber.open(pdf_path) as pdf:
            for pagina in pdf.pages:
                # Layout com grade: consulta de pagamentos.
                for tabela in pagina.extract_tables():
                    for linha in tabela:
                        if len(linha) >= 2 and linha[0] and linha[1]:
                            doc = normalizar_documento(linha[1])
                            if doc:
                                documentos.setdefault(chave_nome(linha[0]), doc)
                # Layout sem grade: extrato de conta corrente.
                for texto in (pagina.extract_text() or "").split("\n"):
                    m = re.match(
                        r"\d{2}/\d{2}/\d{4}\s+(.+?)\s+([\d.]{3,}[-/][\d.\-/]+)\s+-?[\d.]+,\d{2}",
                        texto,
                    )
                    if m:
                        doc = normalizar_documento(m.group(2))
                        if doc:
                            # A descricao bancaria precede a razao social na mesma
                            # linha; guardamos o trecho todo e casamos por prefixo.
                            documentos.setdefault(chave_nome(m.group(1)), doc)
    return documentos


# Abaixo deste tamanho, casar por conteudo gera falso positivo ("ALGCOM",
# "EMBRATEL" aparecem dentro de nomes de terceiros).
MIN_CHAVE_PARA_BUSCA_PARCIAL = 15


def casar_documentos(fornecedores: dict[str, Fornecedor], documentos: dict[str, str]) -> int:
    """Anexa CPF/CNPJ aos fornecedores, por chave exata ou por prefixo longo."""
    casados = 0
    for chave, f in fornecedores.items():
        if chave in documentos:
            f.documentos.add(documentos[chave])
            casados += 1
            continue
        if len(chave) < MIN_CHAVE_PARA_BUSCA_PARCIAL:
            continue
        # O extrato prefixa a razao social com a descricao do lancamento
        # ("BOLETO PAGO EQUATORIAL P EQUATORIAL PARA ..."), entao a chave pode
        # estar no meio. Exigimos correspondencia unica para nao adivinhar.
        achados = {doc for chave_doc, doc in documentos.items() if chave in chave_doc}
        if len(achados) == 1:
            f.documentos.add(achados.pop())
            casados += 1
    return casados


def escrever_csv(caminho: Path, fornecedores: dict[str, Fornecedor]) -> None:
    """Seed do app: uma linha por regra. Fornecedor ambiguo na conta emite N."""
    with caminho.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(
            [
                "documento",
                "nome_canonico",
                "conta_debito",
                "centro_custo",
                "centros_alternativos",
                "confianca",
                "competencias",
            ]
        )
        for f in sorted(fornecedores.values(), key=lambda x: x.nome_canonico):
            confianca = f.classificar()
            for conta in sorted(f.contas):
                meses = sorted(
                    {m for (c, _), ms in f.destinos.items() if c == conta for m in ms}
                )
                w.writerow(
                    [
                        f.documento,
                        f.nome_canonico,
                        conta,
                        f.centro_dominante,
                        ";".join(f.centros_alternativos),
                        confianca,
                        ";".join(meses),
                    ]
                )


def escrever_xlsx(caminho: Path, fornecedores: dict[str, Fornecedor], stats: dict) -> None:
    """Planilha de revisao para o contador: regras prontas, ambiguos, resumo."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Regras"
    ws.append(
        [
            "Documento",
            "Fornecedor",
            "Conta Debito",
            "Centro Custo",
            "Outros centros vistos",
            "Confianca",
            "Meses",
            "Confere? (X)",
        ]
    )
    for f in sorted(fornecedores.values(), key=lambda x: x.nome_canonico):
        if f.classificar() == "AMBIGUO_CONTA":
            continue
        ws.append(
            [
                f.documento,
                f.nome_canonico,
                next(iter(f.contas)),
                f.centro_dominante,
                ", ".join(f.centros_alternativos),
                f.classificar(),
                len(f.competencias),
                "",
            ]
        )

    amb = wb.create_sheet("Ambiguos")
    amb["A1"] = (
        "Estes fornecedores usam mais de uma conta de debito no historico. "
        "Marque com X a conta correta, ou indique o criterio de escolha."
    )
    amb.append([])
    amb.append(["Documento", "Fornecedor", "Conta Debito", "Meses", "Escolher (X)"])
    for f in sorted(fornecedores.values(), key=lambda x: x.nome_canonico):
        if f.classificar() != "AMBIGUO_CONTA":
            continue
        for conta in sorted(f.contas):
            meses = {m for (c, _), ms in f.destinos.items() if c == conta for m in ms}
            amb.append([f.documento, f.nome_canonico, conta, len(meses), ""])
        amb.append([])

    # O CNPJ nao e chave suficiente sozinho: o proprio CNPJ da empresa
    # (19.402.859/0001-55) aparece tanto em transferencia entre contas proprias
    # quanto em conta intercompany. Reportamos para que a Fase 3 resolva por
    # documento + nome, nao por documento apenas.
    por_documento: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for f in fornecedores.values():
        if f.documento:
            for conta in f.contas:
                por_documento[f.documento].add((f.nome_canonico, conta))
    colisoes = {d: v for d, v in por_documento.items() if len({n for n, _ in v}) > 1}
    if colisoes:
        col = wb.create_sheet("Colisoes")
        col["A1"] = (
            "Um mesmo CPF/CNPJ usado por nomes diferentes. A chave de casamento "
            "precisa ser documento + nome, nao documento sozinho."
        )
        col.append([])
        col.append(["Documento", "Fornecedor", "Conta Debito"])
        for doc, pares in sorted(colisoes.items()):
            for nome, conta in sorted(pares):
                col.append([doc, nome, conta])
            col.append([])

    res = wb.create_sheet("Resumo")
    por_classe = defaultdict(int)
    for f in fornecedores.values():
        por_classe[f.classificar()] += 1
    multi_centro = sum(1 for f in fornecedores.values() if f.centros_alternativos)
    res.append(["Metrica", "Valor"])
    for chave, valor in [
        ("Fornecedores distintos", len(fornecedores)),
        ("Confianca ALTA (conta unica, 3+ meses)", por_classe["ALTA"]),
        ("Confianca MEDIA (conta unica, 1-2 meses)", por_classe["MEDIA"]),
        ("AMBIGUO_CONTA (revisar na aba Ambiguos)", por_classe["AMBIGUO_CONTA"]),
        ("Com mais de um centro de custo observado", multi_centro),
        ("Com CPF/CNPJ identificado", sum(1 for f in fornecedores.values() if f.documento)),
        ("Linhas de historico lidas", stats["linhas"]),
        ("Linhas com SISPAG FORNECEDORES", stats["sispag"]),
        ("Linhas nao atribuidas a fornecedor", stats.get("nao_atribuidas", 0)),
        ("Documentos usados por mais de um nome", len(colisoes)),
    ]:
        res.append([chave, valor])
    for nome, origem in stats["duplicatas"]:
        res.append([f"Duplicata ignorada: {nome}", f"identica a {origem}"])

    for planilha in (ws, amb, res):
        for coluna in planilha.columns:
            largura = max((len(str(c.value or "")) for c in coluna), default=10)
            planilha.column_dimensions[coluna[0].column_letter].width = min(largura + 2, 48)

    wb.save(caminho)


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--dir", default="arquivos-clickip/clickIP", type=Path)
    p.add_argument("--out", default="docs", type=Path)
    args = p.parse_args()

    fornecedores, stats = ler_planilhas(args.dir)
    documentos = ler_documentos_itau(args.dir)
    casados = casar_documentos(fornecedores, documentos)

    args.out.mkdir(parents=True, exist_ok=True)
    escrever_csv(args.out / "base-depara-inicial.csv", fornecedores)
    escrever_xlsx(args.out / "base-depara-inicial.xlsx", fornecedores, stats)

    por_classe = defaultdict(int)
    for f in fornecedores.values():
        por_classe[f.classificar()] += 1

    print("Arquivos lidos:")
    for nome, n in stats["arquivos"]:
        print(f"  {nome:40s} {n:4d} linhas")
    for nome, origem in stats["duplicatas"]:
        print(f"  IGNORADO (duplicata de {origem}): {nome}")
    print(f"\nLinhas de historico: {stats['linhas']}  (SISPAG: {stats['sispag']})")
    print(f"Linhas nao atribuidas a fornecedor: {stats.get('nao_atribuidas', 0)}")
    print(f"Documentos CPF/CNPJ extraidos dos PDFs Itau: {len(documentos)}")
    print(f"\nFornecedores distintos: {len(fornecedores)}")
    print(f"  ALTA          {por_classe['ALTA']:4d}  conta unica, 3+ meses")
    print(f"  MEDIA         {por_classe['MEDIA']:4d}  conta unica, 1-2 meses")
    print(f"  AMBIGUO_CONTA {por_classe['AMBIGUO_CONTA']:4d}  revisao humana obrigatoria")
    print(f"  com CPF/CNPJ  {casados:4d}")
    print(
        f"  com >1 centro de custo "
        f"{sum(1 for f in fornecedores.values() if f.centros_alternativos):4d}  "
        f"centro proposto = dominante, ajustavel na validacao"
    )
    print(f"\nEscrito em {args.out}/base-depara-inicial.{{csv,xlsx}}")


if __name__ == "__main__":
    main()
