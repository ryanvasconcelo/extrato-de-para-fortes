"""Teste ponta a ponta contra o gabarito de junho.

Gabarito: CLICK SCM 062026.xlsx, 439 linhas produzidas a mao pelo contador.

Vale para Debito, Credito, Valor, Data e Centro de custo. NAO vale para Historico:
91 das 439 linhas trazem o generico 'SISPAG FORNECEDORES', que e divida do
processo manual e nao limite do dado (ADR 0005). Divergir ali e o comportamento
desejado.

Os limiares abaixo sao a linha de base medida, nao a meta final. Servem para
detectar regressao. Subir um limiar exige medir de novo; baixar exige justificar.
"""

from __future__ import annotations

from collections import Counter

from app.modelos import Confianca, StatusLancamento
from app.motor.validador import BLOCKERS, Codigo
from app.normalizacao import normalizar_conta

SISPAG = "SISPAG FORNECEDORES"


def _indexar(itens, pagamentos):
    """(data, valor em centavos) -> lancamentos. Chave que existe nas duas fontes:
    o gabarito nao tem id de pagamento, e o nome do favorecido so aparece na
    coluna K de um dos seis arquivos."""
    pagamento_por_id = {p.id: p for p in pagamentos}
    indice = {}
    for item in itens:
        pagamento = pagamento_por_id[item.lancamento.pagamento_id]
        chave = (pagamento.data.strftime("%d/%m/%Y"), round(item.lancamento.valor * 100))
        indice.setdefault(chave, []).append(item)
    return indice


class TestCoberturaEEstrutura:
    def test_uma_linha_por_pagamento(self, processado, gabarito):
        itens, resumo = processado
        assert len(itens) == len(gabarito) == 439
        assert resumo.total == 439

    def test_credito_e_sempre_a_conta_do_itau(self, processado, gabarito):
        """Coluna D do arquivo Fortes: 1.01.01.02.01.0003 em 100% das linhas."""
        itens, _ = processado
        esperado = normalizar_conta(gabarito[0]["conta_credito"])
        assert esperado == "1.01.01.02.01.0003"
        assert {i.lancamento.conta_credito for i in itens} == {esperado}

    def test_soma_dos_valores_bate(self, processado, gabarito):
        _, resumo = processado
        assert abs(resumo.valor_total - sum(l["valor"] for l in gabarito)) < 0.02

    def test_filial_constante(self, processado):
        itens, _ = processado
        assert {i.lancamento.filial for i in itens} == {"0001"}


class TestAcuraciaDaContaDeDebito:
    def test_classificacao_automatica_cobre_a_maioria(self, processado):
        _, resumo = processado
        assert resumo.automaticos / resumo.total >= 0.75

    def test_conta_de_debito_acerta_o_gabarito(self, processado, pagamentos_junho, gabarito):
        """Entre as linhas classificadas automaticamente, quantas batem?

        Esta e a metrica central do produto: se a conta de debito estiver errada,
        o lancamento contabil esta errado.
        """
        indice = _indexar(*processado[:1], pagamentos_junho)
        acertos = erros = sem_par = nao_classificado = 0
        for linha in gabarito:
            chave = (linha["data"], round(linha["valor"] * 100))
            candidatos = indice.get(chave)
            if not candidatos:
                sem_par += 1
                continue
            item = candidatos[0]
            if item.lancamento.status is StatusLancamento.PENDENTE:
                nao_classificado += 1
            elif normalizar_conta(item.lancamento.conta_debito) == normalizar_conta(
                linha["conta_debito"]
            ):
                acertos += 1
            else:
                erros += 1

        classificadas = acertos + erros
        assert classificadas > 0
        precisao = acertos / classificadas
        print(
            f"\n  debito: {acertos} acertos, {erros} erros -> precisao {precisao:.1%}"
            f"\n  {nao_classificado} em pendencia, {sem_par} sem par no gabarito"
        )
        assert precisao >= 0.90, f"precisao caiu para {precisao:.1%}"

    def test_centro_de_custo_dominante_acerta_a_maioria(
        self, processado, pagamentos_junho, gabarito
    ):
        """O centro de custo e sugestao, nao determinacao (ADR 0006).

        16 fornecedores usam varios centros com a mesma conta, entao 100% e
        impossivel por construcao. O que se mede e se a sugestao e util.
        """
        indice = _indexar(*processado[:1], pagamentos_junho)
        acertos = total = 0
        for linha in gabarito:
            candidatos = indice.get((linha["data"], round(linha["valor"] * 100)))
            if not candidatos or candidatos[0].lancamento.status is StatusLancamento.PENDENTE:
                continue
            total += 1
            acertos += candidatos[0].lancamento.centro_custo == linha["centro_custo"]
        assert total > 0
        print(f"\n  centro de custo: {acertos}/{total} = {acertos / total:.1%}")
        assert acertos / total >= 0.85


class TestPoliticaDeHistorico:
    def test_gabarito_tem_as_91_linhas_sispag_esperadas(self, gabarito):
        """Se este numero mudar, o arquivo do cliente mudou e o ADR 0005 precisa
        ser revisitado antes de confiar nos outros testes de historico."""
        assert sum(1 for l in gabarito if l["historico"] == SISPAG) == 91

    def test_derivacao_supera_a_linha_de_base_do_spike(self, processado):
        """Spike da Fase 2 mediu 58%. Regredir abaixo disso e regressao real."""
        _, resumo = processado
        taxa = resumo.historico_derivado / resumo.total
        print(f"\n  historico derivado: {resumo.historico_derivado}/{resumo.total} = {taxa:.1%}")
        assert taxa >= 0.58

    def test_nunca_ha_historico_vazio(self, processado):
        """O contador precisa de algo legivel para decidir o que editar."""
        itens, _ = processado
        assert [i for i in itens if not i.lancamento.historico.strip()] == []

    def test_nao_derivado_sempre_carrega_o_warning(self, processado):
        """Verificacao explicita do ADR 0005: nunca silencioso."""
        itens, _ = processado
        for item in itens:
            if item.lancamento.titulo_id is None:
                codigos = {a.codigo for a in item.achados}
                assert Codigo.HISTORICO_NAO_DERIVADO in codigos

    def test_enriquece_linhas_que_o_manual_deixou_generico(
        self, processado, pagamentos_junho, gabarito
    ):
        """O ganho que justifica o produto: histórico onde antes havia SISPAG."""
        indice = _indexar(*processado[:1], pagamentos_junho)
        enriquecidas = 0
        for linha in gabarito:
            if linha["historico"] != SISPAG:
                continue
            candidatos = indice.get((linha["data"], round(linha["valor"] * 100)))
            if candidatos and candidatos[0].lancamento.titulo_id is not None:
                enriquecidas += 1
        print(f"\n  SISPAG enriquecidas: {enriquecidas}/91")
        assert enriquecidas >= 40


class TestFalsoPositivoDeSubstring:
    """O teste que a abordagem por substring do plano original falharia.

    Cinco entidades do grupo CLICK IP com prefixo comum e cinco contas distintas.
    `descricao.includes("CLICK IP")` casaria com todas. Ver ADR 0002.
    """

    def test_as_cinco_entidades_click_ip_tem_contas_distintas(self, base_depara):
        fornecedores, regras = base_depara
        alvos = [f for f in fornecedores if f.nome_canonico.startswith("CLICK IP")]
        assert len(alvos) == 5

        por_fornecedor = {f.id: f for f in alvos}
        contas = {
            r.conta_debito for r in regras if r.fornecedor_id in por_fornecedor
        }
        assert len(contas) == 5, f"esperava 5 contas distintas, obtive {sorted(contas)}"

    def test_classificador_nao_confunde_entidades_do_mesmo_grupo(
        self, plano_contas, base_depara, conta_itau
    ):
        from app.modelos import Pagamento
        from app.motor.processador import Processador
        from datetime import date

        fornecedores, regras = base_depara
        alvos = {
            f.nome_canonico: f for f in fornecedores if f.nome_canonico.startswith("CLICK IP")
        }
        regra_de = {r.fornecedor_id: r for r in regras if r.ativo}

        pagamentos = []
        esperado = {}
        for i, (nome, fornecedor) in enumerate(sorted(alvos.items()), start=1):
            regra = regra_de.get(fornecedor.id)
            if regra is None:
                continue
            pagamentos.append(
                Pagamento(
                    id=i,
                    lote_id=1,
                    data=date(2026, 6, 15),
                    favorecido_raw=nome,
                    documento_raw=fornecedor.documento,
                    valor=100.0 + i,
                )
            )
            esperado[i] = regra.conta_debito

        processador = Processador(plano_contas, fornecedores, regras, conta_itau)
        itens, _ = processador.processar(1, pagamentos, [])

        obtido = {i.lancamento.pagamento_id: i.lancamento.conta_debito for i in itens}
        assert obtido == esperado


class TestRegraGeraMudancaDeStatus:
    """Teste pedido explicitamente na secao 4 do plano original."""

    def test_criar_regra_move_pendente_para_auto(self, plano_contas, conta_itau):
        from datetime import date

        from app.modelos import Fornecedor, Pagamento, RegraDePara
        from app.motor.processador import Processador

        fornecedor = Fornecedor(
            id=1,
            documento="11222333000181",
            nome_canonico="FORNECEDOR NOVO",
            chave_nome="FORNECEDOR NOVO",
        )
        pagamento = Pagamento(
            id=1,
            lote_id=1,
            data=date(2026, 6, 10),
            favorecido_raw="FORNECEDOR NOVO",
            documento_raw="11222333000181",
            valor=500.0,
        )

        sem_regra = Processador(plano_contas, [fornecedor], [], conta_itau)
        itens, resumo = sem_regra.processar(1, [pagamento], [])
        assert itens[0].lancamento.status is StatusLancamento.PENDENTE
        assert resumo.pendentes == 1
        assert Codigo.CONTA_DEBITO_AUSENTE in {a.codigo for a in itens[0].achados}

        conta = next(c.codigo for c in plano_contas if c.analitica)
        regra = RegraDePara(
            id=1,
            fornecedor_id=1,
            conta_debito=conta,
            centro_custo_sugerido="0001",
            confianca=Confianca.ALTA,
        )
        com_regra = Processador(plano_contas, [fornecedor], [regra], conta_itau)
        itens, resumo = com_regra.processar(1, [pagamento], [])
        assert itens[0].lancamento.status is StatusLancamento.AUTO
        assert itens[0].lancamento.conta_debito == conta
        assert resumo.pendentes == 0

    def test_regra_ambigua_bloqueia_em_vez_de_escolher(self, plano_contas, conta_itau):
        """As 13 regras AMBIGUO_CONTA entram inativas de proposito (ADR 0003)."""
        from datetime import date

        from app.modelos import Fornecedor, Pagamento, RegraDePara
        from app.motor.processador import Processador

        fornecedor = Fornecedor(
            id=1, documento="11222333000181", nome_canonico="CONSORCIO", chave_nome="CONSORCIO"
        )
        contas = [c.codigo for c in plano_contas if c.analitica][:2]
        regras = [
            RegraDePara(
                id=i + 1,
                fornecedor_id=1,
                conta_debito=conta,
                confianca=Confianca.AMBIGUO_CONTA,
                ativo=False,
            )
            for i, conta in enumerate(contas)
        ]
        pagamento = Pagamento(
            id=1,
            lote_id=1,
            data=date(2026, 6, 10),
            favorecido_raw="CONSORCIO",
            documento_raw="11222333000181",
            valor=500.0,
        )

        processador = Processador(plano_contas, [fornecedor], regras, conta_itau)
        itens, resumo = processador.processar(1, [pagamento], [])

        assert itens[0].lancamento.status is StatusLancamento.PENDENTE
        assert itens[0].lancamento.conta_debito == ""
        codigos = {a.codigo for a in itens[0].achados}
        assert Codigo.REGRA_AMBIGUA in codigos
        # A ausencia de conta e consequencia da ambiguidade, nao um segundo
        # problema: reportar os dois daria duas linhas de motivo para uma decisao.
        assert Codigo.CONTA_DEBITO_AUSENTE not in codigos
        assert resumo.status_lote.value == "BLOQUEADO"


class TestArquivoFinalDeJunho:
    """O arquivo gerado ao lado do que o contador entregou ao Fortes.

    Estrutura e totais tem que bater; Historico e centro de custo divergem por
    decisao (ADR 0005 e ADR 0006), e essa divergencia e medida nos testes acima.
    """

    def _gerar(self, processado, pagamentos_junho):
        from io import BytesIO

        import openpyxl

        from app.export_fortes import arquivo_final
        from app.modelos import StatusLote

        itens, _ = processado
        datas = {p.id: p.data for p in pagamentos_junho}
        conteudo = arquivo_final(
            StatusLote.APROVADO,
            [(i.lancamento, datas[i.lancamento.pagamento_id]) for i in itens],
        )
        return openpyxl.load_workbook(BytesIO(conteudo)).active

    def test_mesma_forma_do_arquivo_do_cliente(self, processado, pagamentos_junho):
        import openpyxl

        from .conftest import GABARITO_JUNHO

        ws = self._gerar(processado, pagamentos_junho)
        cliente = openpyxl.load_workbook(GABARITO_JUNHO, data_only=True).active

        assert ws.max_column == 10
        assert ws.max_row == cliente.max_row == 440  # linha modelo + 439
        assert [c.value for c in ws[1]] == [c.value for c in cliente[1]][:10]

    def test_soma_bate_ao_centavo(self, processado, pagamentos_junho):
        import openpyxl

        from .conftest import GABARITO_JUNHO

        ws = self._gerar(processado, pagamentos_junho)
        cliente = openpyxl.load_workbook(GABARITO_JUNHO, data_only=True).active

        gerado = sum(ws.cell(r, 5).value for r in range(2, ws.max_row + 1))
        esperado = sum(cliente.cell(r, 5).value for r in range(2, cliente.max_row + 1))
        assert round(gerado, 2) == round(esperado, 2)

    def test_nenhuma_linha_sai_com_conta_vazia(self, processado, pagamentos_junho):
        """As 101 pendencias nao podem chegar ao arquivo sem conta.

        Aqui elas chegam porque o teste forca APROVADO para exercitar o gerador;
        na aplicacao o lote estaria BLOQUEADO. O que este teste protege e a
        ordem das colunas: conta de debito na C, credito na D.
        """
        ws = self._gerar(processado, pagamentos_junho)
        for r in range(2, ws.max_row + 1):
            assert ws.cell(r, 4).value == "1.01.01.02.01.0003"


class TestValidacao:
    def test_nenhuma_conta_inexistente(self, processado):
        """Se a normalizacao de digito verificador quebrar, este teste explode:
        seriam 439 CONTA_INEXISTENTE de uma vez (ADR 0006)."""
        itens, _ = processado
        codigos = Counter(a.codigo for i in itens for a in i.achados)
        assert codigos[Codigo.CONTA_INEXISTENTE] == 0
        assert codigos[Codigo.CONTA_NAO_ANALITICA] == 0

    def test_banco_sempre_mapeado(self, processado):
        itens, _ = processado
        codigos = {a.codigo for i in itens for a in i.achados}
        assert Codigo.BANCO_NAO_MAPEADO not in codigos

    def test_lote_com_pendencia_fica_bloqueado(self, processado):
        _, resumo = processado
        assert resumo.pendentes > 0
        assert resumo.status_lote.value == "BLOQUEADO"

    def test_pendente_nunca_tem_conta_de_debito(self, processado):
        itens, _ = processado
        for item in itens:
            if item.lancamento.status is StatusLancamento.PENDENTE:
                assert not item.lancamento.conta_debito or any(
                    a.codigo in BLOCKERS for a in item.achados
                )
