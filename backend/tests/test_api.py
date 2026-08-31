"""Testes da API sobre os PDFs reais do cliente.

O que o teste de motor (test_ponta_a_ponta.py) nao cobre e o que so existe na
fronteira HTTP: deteccao de layout no upload, reimportacao do mesmo arquivo,
persistencia do reprocessamento, e a trava de exportacao.

O banco e um SQLite temporario por sessao de teste. Semear le o plano de contas
real e a base De/Para minerada, entao a fixture custa alguns segundos - por isso
escopo de sessao e um lote unico compartilhado.
"""

from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient

from .conftest import CONTAS_PAGAR, ITAU_EXTRATO, ITAU_PAGAMENTOS


@pytest.fixture(scope="module")
def cliente(tmp_path_factory, monkeypatch_sessao):
    """TestClient com banco proprio.

    DATABASE_PATH tem que ser definido ANTES de importar app.banco, porque o
    engine e criado no import. Dai o monkeypatch de escopo de sessao.
    """
    destino = tmp_path_factory.mktemp("banco") / "teste.db"
    monkeypatch_sessao.setenv("DATABASE_PATH", str(destino))

    from app.api import app

    with TestClient(app) as c:
        yield c


@pytest.fixture(scope="module")
def monkeypatch_sessao():
    from _pytest.monkeypatch import MonkeyPatch

    mp = MonkeyPatch()
    yield mp
    mp.undo()


@pytest.fixture(scope="module")
def lote_junho(cliente):
    """Junho importado pelas tres vias: pagamentos, extrato e Contas a Pagar."""
    lote = cliente.post("/api/lotes", params={"competencia": "062026"}).json()
    for caminho in (ITAU_PAGAMENTOS, ITAU_EXTRATO, CONTAS_PAGAR):
        resposta = cliente.post(
            f"/api/lotes/{lote['id']}/arquivos",
            files={"arquivo": (caminho.name, caminho.read_bytes(), "application/pdf")},
        )
        assert resposta.status_code == 200, resposta.text
    return lote


class TestSeed:
    def test_semeia_plano_e_base_depara_no_startup(self, cliente):
        saude = cliente.get("/api/saude").json()
        # 1.750 contas, nao as 1.864 que os documentos da Fase 0 e 2 registram:
        # aquele numero contava linha de planilha, inclusive cabecalho e branco.
        # Ver ADR 0009, secao de correcoes de medicao.
        assert saude["plano_contas"] == 1750
        assert saude["regras"] > 150
        # As AMBIGUO_CONTA entram inativas de proposito (ADR 0003).
        assert saude["regras_ativas"] < saude["regras"]


class TestImportacao:
    def test_detecta_os_tres_layouts(self, cliente, lote_junho):
        del lote_junho  # a fixture ja assertou o 200 de cada upload
        lote = cliente.get("/api/lotes").json()[0]
        assert lote["lancamentos"] == 439

    def test_reimportar_o_mesmo_arquivo_e_recusado(self, cliente, lote_junho):
        """RF-01.8: reimportar duplicaria 439 linhas em silencio."""
        resposta = cliente.post(
            f"/api/lotes/{lote_junho['id']}/arquivos",
            files={
                "arquivo": (
                    ITAU_PAGAMENTOS.name,
                    ITAU_PAGAMENTOS.read_bytes(),
                    "application/pdf",
                )
            },
        )
        assert resposta.status_code == 409
        assert "ja importado" in resposta.json()["detail"]

    def test_xlsx_e_recusado_com_explicacao(self, cliente, lote_junho):
        resposta = cliente.post(
            f"/api/lotes/{lote_junho['id']}/arquivos",
            files={"arquivo": ("plano.xlsx", b"nao importa", "application/vnd.ms-excel")},
        )
        assert resposta.status_code == 415

    def test_conta_de_credito_vem_do_cabecalho(self, cliente, lote_junho):
        """A conta de credito nao e constante no codigo: sai do cabecalho do
        relatorio, casada contra a Base Bancos (handover FASE-1)."""
        linhas = cliente.get(f"/api/lotes/{lote_junho['id']}/lancamentos").json()
        assert {l["conta_credito"] for l in linhas} == {"1.01.01.02.01.0003"}


class TestPendenciasAgrupadas:
    def test_agrupa_por_fornecedor(self, cliente, lote_junho):
        pendencias = cliente.get(f"/api/lotes/{lote_junho['id']}/pendencias").json()
        assert pendencias
        # Agrupado significa menos grupos do que linhas.
        assert len(pendencias) < sum(p["linhas"] for p in pendencias)
        assert pendencias == sorted(pendencias, key=lambda p: -p["linhas"])

    def test_toda_pendencia_tem_motivo_e_mensagem(self, cliente, lote_junho):
        for p in cliente.get(f"/api/lotes/{lote_junho['id']}/pendencias").json():
            assert p["motivos"]
            assert p["mensagens"]

    def test_ambiguidade_nao_duplica_motivo(self, cliente, lote_junho):
        """Fornecedor ambiguo mostra a ambiguidade, nao tambem a conta ausente."""
        for p in cliente.get(f"/api/lotes/{lote_junho['id']}/pendencias").json():
            if "REGRA_AMBIGUA" in p["motivos"]:
                assert p["motivos"] == ["REGRA_AMBIGUA"]

    def test_reprocessar_nao_muda_resultado_sem_regra_nova(self, cliente, lote_junho):
        """Idempotencia: reprocessar duas vezes seguidas da o mesmo lote."""
        primeiro = cliente.post(f"/api/lotes/{lote_junho['id']}/reprocessar").json()
        segundo = cliente.post(f"/api/lotes/{lote_junho['id']}/reprocessar").json()
        assert primeiro == segundo
        assert primeiro["total"] == 439


class TestRegraMudaStatusPelaApi:
    """Mesmo teste do plano original, agora atravessando HTTP e SQLite.

    O de test_ponta_a_ponta prova que o motor muda de status; este prova que a
    mudanca sobrevive ao reprocessamento e a persistencia.
    """

    def test_criar_regra_resolve_todas_as_linhas_do_fornecedor(self, cliente, lote_junho):
        pendencias = cliente.get(f"/api/lotes/{lote_junho['id']}/pendencias").json()
        alvo = next(p for p in pendencias if p["linhas"] >= 2)
        conta = cliente.get("/api/plano-contas", params={"q": "3."}).json()[0]["codigo"]

        antes = cliente.get(
            f"/api/lotes/{lote_junho['id']}/lancamentos", params={"status": "PENDENTE"}
        ).json()

        resposta = cliente.post(
            "/api/regras",
            json={
                "fornecedor_nome": alvo["fornecedor"],
                "documento": alvo["documento"],
                "conta_debito": conta,
                "centro_custo": "0001",
            },
        )
        assert resposta.status_code == 200, resposta.text

        depois = cliente.get(
            f"/api/lotes/{lote_junho['id']}/lancamentos", params={"status": "PENDENTE"}
        ).json()
        assert len(depois) == len(antes) - alvo["linhas"]

        automaticos = cliente.get(
            f"/api/lotes/{lote_junho['id']}/lancamentos", params={"status": "AUTO"}
        ).json()
        assert any(l["conta_debito"] == conta for l in automaticos)

    def test_conta_fora_do_plano_e_recusada(self, cliente, lote_junho):
        linha = cliente.get(f"/api/lotes/{lote_junho['id']}/lancamentos").json()[0]
        resposta = cliente.patch(
            f"/api/lancamentos/{linha['id']}", json={"conta_debito": "9.99.99.99.99.9999"}
        )
        assert resposta.status_code == 422

    def test_edicao_manual_sobrevive_ao_reprocessamento(self, cliente, lote_junho):
        """RF-05.2: reprocessar recalcula tudo, menos o que o humano decidiu."""
        linha = cliente.get(
            f"/api/lotes/{lote_junho['id']}/lancamentos", params={"status": "AUTO"}
        ).json()[0]
        conta = cliente.get("/api/plano-contas", params={"q": "4."}).json()[0]["codigo"]

        valor_novo = round(linha["valor"] + 1.0, 2)
        credito_novo = next(
            c["codigo"]
            for c in cliente.get("/api/plano-contas", params={"q": "1.01.01"}).json()
            if c["codigo"] != linha["conta_credito"]
        )
        cliente.patch(
            f"/api/lancamentos/{linha['id']}",
            json={
                "conta_debito": conta,
                "conta_credito": credito_novo,
                "valor": valor_novo,
                "historico": "AJUSTE MANUAL DE CONFERENCIA",
                "criar_regra": False,
            },
        )
        # Uma regra qualquer forca o reprocessamento do lote inteiro.
        cliente.post(
            "/api/regras",
            json={
                "fornecedor_nome": "FORNECEDOR INVENTADO PARA REPROCESSAR",
                "documento": "",
                "conta_debito": conta,
                "centro_custo": "0001",
            },
        )

        depois = cliente.get(f"/api/lotes/{lote_junho['id']}/lancamentos").json()
        editada = next(
            l for l in depois if l["historico"] == "AJUSTE MANUAL DE CONFERENCIA"
        )
        assert editada["conta_debito"] == conta
        assert editada["conta_credito"] == credito_novo
        assert editada["valor"] == valor_novo
        assert editada["status"] == "MANUAL"


class TestExcecaoPorLinha:
    def _par_auto(self, cliente, lote_id):
        from collections import defaultdict

        auto = cliente.get(
            f"/api/lotes/{lote_id}/lancamentos", params={"status": "AUTO"}
        ).json()
        por_doc = defaultdict(list)
        for linha in auto:
            if linha["documento"]:
                por_doc[linha["documento"]].append(linha)
        par = next(v for v in por_doc.values() if len(v) >= 2)
        return par[0], par[1]

    def test_editar_uma_linha_nao_mexe_na_irma(self, cliente, lote_junho):
        a, b = self._par_auto(cliente, lote_junho["id"])
        id_a, id_b = a["id"], b["id"]
        conta_b = b["conta_debito"]
        conta_nova = cliente.get("/api/plano-contas", params={"q": "4."}).json()[0][
            "codigo"
        ]
        assert conta_nova != a["conta_debito"]

        resposta = cliente.patch(
            f"/api/lancamentos/{id_a}",
            json={"conta_debito": conta_nova, "criar_regra": False},
        )
        assert resposta.status_code == 200

        depois = {
            l["id"]: l
            for l in cliente.get(f"/api/lotes/{lote_junho['id']}/lancamentos").json()
        }
        assert id_a in depois and id_b in depois, (
            "edicao sem regra nao pode recriar o lote e trocar os ids"
        )
        assert depois[id_a]["conta_debito"] == conta_nova
        assert depois[id_a]["status"] == "MANUAL"
        assert depois[id_b]["conta_debito"] == conta_b
        assert depois[id_b]["status"] == "AUTO"

    def test_editar_valor_e_credito_nao_mexe_na_irma(self, cliente, lote_junho):
        a, b = self._par_auto(cliente, lote_junho["id"])
        id_a, id_b = a["id"], b["id"]
        valor_b, credito_b = b["valor"], b["conta_credito"]
        valor_novo = round(a["valor"] + 10.0, 2)
        credito_novo = next(
            c["codigo"]
            for c in cliente.get("/api/plano-contas", params={"q": "1.01.01"}).json()
            if c["codigo"] != a["conta_credito"]
        )

        resposta = cliente.patch(
            f"/api/lancamentos/{id_a}",
            json={
                "conta_credito": credito_novo,
                "valor": valor_novo,
                "criar_regra": False,
            },
        )
        assert resposta.status_code == 200

        depois = {
            l["id"]: l
            for l in cliente.get(f"/api/lotes/{lote_junho['id']}/lancamentos").json()
        }
        assert id_a in depois and id_b in depois, (
            "edicao sem regra nao pode recriar o lote e trocar os ids"
        )
        assert depois[id_a]["valor"] == valor_novo
        assert depois[id_a]["conta_credito"] == credito_novo
        assert depois[id_a]["status"] == "MANUAL"
        assert depois[id_b]["valor"] == valor_b
        assert depois[id_b]["conta_credito"] == credito_b
        assert depois[id_b]["status"] == "AUTO"

    def test_valor_nao_positivo_e_recusado(self, cliente, lote_junho):
        a, _ = self._par_auto(cliente, lote_junho["id"])
        resposta = cliente.patch(
            f"/api/lancamentos/{a['id']}",
            json={"valor": 0, "criar_regra": False},
        )
        assert resposta.status_code == 422
        assert "Valor nao e positivo" in resposta.json()["detail"]

    def test_criar_regra_false_nao_chama_reprocessar(
        self, cliente, lote_junho, monkeypatch
    ):
        def _explodir(*_a, **_k):
            raise AssertionError("criar_regra false nao pode reprocessar o lote")

        monkeypatch.setattr("app.api._reprocessar", _explodir)
        a, _ = self._par_auto(cliente, lote_junho["id"])
        conta = cliente.get("/api/plano-contas", params={"q": "4."}).json()[0]["codigo"]
        resposta = cliente.patch(
            f"/api/lancamentos/{a['id']}",
            json={"conta_debito": conta, "criar_regra": False},
        )
        assert resposta.status_code == 200

    def test_edicao_preserva_avisos_do_classificador(self, cliente, lote_junho):
        avisos_classificador = {
            "HISTORICO_NAO_DERIVADO",
            "FORNECEDOR_SEM_DOCUMENTO",
            "REGRA_CONFIANCA_MEDIA",
        }
        alvo = next(
            l
            for l in cliente.get(f"/api/lotes/{lote_junho['id']}/lancamentos").json()
            if any(w in avisos_classificador for w in l["warnings"])
        )
        avisos = [w for w in alvo["warnings"] if w in avisos_classificador]
        conta = cliente.get("/api/plano-contas", params={"q": "4."}).json()[0]["codigo"]
        resposta = cliente.patch(
            f"/api/lancamentos/{alvo['id']}",
            json={"conta_debito": conta, "criar_regra": False},
        )
        assert resposta.status_code == 200
        depois = next(
            l
            for l in cliente.get(f"/api/lotes/{lote_junho['id']}/lancamentos").json()
            if l["id"] == alvo["id"]
        )
        assert set(avisos) <= set(depois["warnings"])

    def test_criar_regra_true_ainda_reprocessa_irmas(self, cliente, lote_junho):
        a, b = self._par_auto(cliente, lote_junho["id"])
        documento = a["documento"]
        conta_nova = cliente.get("/api/plano-contas", params={"q": "3.01"}).json()[0][
            "codigo"
        ]
        cliente.patch(
            f"/api/lancamentos/{a['id']}",
            json={"conta_debito": conta_nova, "criar_regra": True},
        )
        depois = cliente.get(f"/api/lotes/{lote_junho['id']}/lancamentos").json()
        irmas_auto = [
            l
            for l in depois
            if l["documento"] == documento and l["status"] == "AUTO"
        ]
        assert irmas_auto, "criar_regra deve reclassificar as irmas como AUTO"
        assert all(l["conta_debito"] == conta_nova for l in irmas_auto)


class TestTravaDeExportacao:
    def test_conferencia_sai_antes_da_aprovacao(self, cliente, lote_junho):
        """Excel de conferencia primeiro, arquivo final depois (RF-06.3)."""
        resposta = cliente.get(f"/api/lotes/{lote_junho['id']}/conferencia")
        assert resposta.status_code == 200
        ws = openpyxl.load_workbook(BytesIO(resposta.content)).active
        assert ws.max_row == 440  # 439 linhas + cabecalho
        assert ws.cell(1, 9).value == "Ocorrencias"

    def test_lote_bloqueado_nao_aprova(self, cliente, lote_junho):
        resposta = cliente.post(f"/api/lotes/{lote_junho['id']}/aprovar")
        assert resposta.status_code == 409
        assert "impedimentos" in resposta.json()["detail"]

    def test_lote_bloqueado_nao_exporta(self, cliente, lote_junho):
        resposta = cliente.get(f"/api/lotes/{lote_junho['id']}/exportar")
        assert resposta.status_code == 409
        assert "aprovacao humana" in resposta.json()["detail"]


class TestExportacaoLiberada:
    """Lote separado, minimo, para exercitar aprovacao e export sem depender de
    resolver as ~90 pendencias reais de junho."""

    @pytest.fixture(scope="class")
    def lote_limpo(self, cliente):
        lote = cliente.post("/api/lotes", params={"competencia": "072026"}).json()
        cliente.post(
            f"/api/lotes/{lote['id']}/arquivos",
            files={
                "arquivo": (
                    ITAU_EXTRATO.name,
                    ITAU_EXTRATO.read_bytes(),
                    "application/pdf",
                )
            },
        )
        # Resolver tudo: cada pendencia vira regra.
        conta = cliente.get("/api/plano-contas", params={"q": "3."}).json()[0]["codigo"]
        for p in cliente.get(f"/api/lotes/{lote['id']}/pendencias").json():
            cliente.post(
                "/api/regras",
                json={
                    "fornecedor_nome": p["fornecedor"],
                    "documento": p["documento"],
                    "conta_debito": conta,
                    "centro_custo": "0001",
                },
            )
        return lote

    @pytest.fixture
    def lote_exportavel(self, cliente):
        lote = cliente.post("/api/lotes", params={"competencia": "082026"}).json()
        cliente.post(
            f"/api/lotes/{lote['id']}/arquivos",
            files={
                "arquivo": (
                    ITAU_EXTRATO.name,
                    ITAU_EXTRATO.read_bytes(),
                    "application/pdf",
                )
            },
        )
        conta = cliente.get("/api/plano-contas", params={"q": "3."}).json()[0]["codigo"]
        for p in cliente.get(f"/api/lotes/{lote['id']}/pendencias").json():
            cliente.post(
                "/api/regras",
                json={
                    "fornecedor_nome": p["fornecedor"],
                    "documento": p["documento"],
                    "conta_debito": conta,
                    "centro_custo": "0001",
                },
            )
        return lote

    def test_sem_pendencia_o_lote_fica_pronto(self, cliente, lote_limpo):
        restantes = cliente.get(f"/api/lotes/{lote_limpo['id']}/pendencias").json()
        assert restantes == []
        lote = next(
            l for l in cliente.get("/api/lotes").json() if l["id"] == lote_limpo["id"]
        )
        assert lote["status"] == "PRONTO"

    def test_aprovar_e_exportar_reproduz_o_formato_do_cliente(self, cliente, lote_limpo):
        assert cliente.post(f"/api/lotes/{lote_limpo['id']}/aprovar").status_code == 200

        resposta = cliente.get(f"/api/lotes/{lote_limpo['id']}/exportar")
        assert resposta.status_code == 200
        ws = openpyxl.load_workbook(BytesIO(resposta.content)).active

        assert ws.max_column == 10

        # Linha 1 e o modelo hibrido dos seis arquivos do cliente: rotulo nas
        # colunas variaveis, constante nas fixas. Ver ADR 0010.
        assert [c.value for c in ws[1]] == [
            "0001",
            "Data",
            "Débito",
            "Crédito",
            " Valor ",
            "Histórico",
            "0001",
            "001",
            "0001",
            "001",
        ]

        primeira = [c.value for c in ws[2]]
        assert primeira[0] == "0001"
        assert primeira[3] == "1.01.01.02.01.0003"
        assert primeira[-3:] == ["001", "0001", "001"]
        # Numero, nao texto: valor como string faz o Fortes recusar o lote.
        assert isinstance(primeira[4], (int, float))

        datas = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1)]
        assert datas == sorted(datas, key=lambda d: d[6:] + d[3:5] + d[:2])

    def test_linha_modelo_confere_com_o_arquivo_do_cliente(self):
        """Compara contra a fonte, nao contra uma copia no teste.

        Se o cliente mudar o modelo, este teste falha e o ADR 0010 precisa de
        sucessor - o layout foi derivado empiricamente por falta do RCO010.
        """
        from app.export_fortes import LINHA_MODELO
        from .conftest import GABARITO_JUNHO

        ws = openpyxl.load_workbook(GABARITO_JUNHO, data_only=True).active
        do_cliente = [c.value for c in ws[1]][: len(LINHA_MODELO)]
        assert LINHA_MODELO == do_cliente

    def test_exportado_nao_aceita_mais_edicao(self, cliente, lote_limpo):
        linha = cliente.get(f"/api/lotes/{lote_limpo['id']}/lancamentos").json()[0]
        assert (
            cliente.patch(
                f"/api/lancamentos/{linha['id']}", json={"centro_custo": "9999"}
            ).status_code
            == 409
        )
        assert (
            cliente.post(
                f"/api/lotes/{lote_limpo['id']}/arquivos",
                files={"arquivo": ("x.pdf", b"x", "application/pdf")},
            ).status_code
            == 409
        )

    def test_edicao_manual_sai_no_arquivo_final(self, cliente, lote_exportavel):
        linhas = cliente.get(f"/api/lotes/{lote_exportavel['id']}/lancamentos").json()
        alvo = linhas[0]
        conta = cliente.get("/api/plano-contas", params={"q": "2.01"}).json()[0]["codigo"]
        cliente.patch(
            f"/api/lancamentos/{alvo['id']}",
            json={
                "conta_debito": conta,
                "historico": "EXCECAO POR LINHA NO EXPORT",
                "centro_custo": "0009",
                "valor": 123.45,
                "criar_regra": False,
            },
        )
        assert cliente.post(f"/api/lotes/{lote_exportavel['id']}/aprovar").status_code == 200
        resposta = cliente.get(f"/api/lotes/{lote_exportavel['id']}/exportar")
        assert resposta.status_code == 200
        ws = openpyxl.load_workbook(BytesIO(resposta.content)).active
        encontrados = [
            row
            for row in ws.iter_rows(min_row=2, max_col=10, values_only=True)
            if row[5] == "EXCECAO POR LINHA NO EXPORT"
        ]
        assert len(encontrados) == 1
        linha = encontrados[0]
        assert linha[2] == conta
        assert linha[4] == 123.45
        assert linha[6] == "0009"


class TestAdministracaoDeCompetencias:
    def test_listar_lotes_inclui_dois_meses(self, cliente):
        a = cliente.post("/api/lotes", params={"competencia": "012027"}).json()
        b = cliente.post("/api/lotes", params={"competencia": "022027"}).json()
        assert a["id"] != b["id"]
        lista = cliente.get("/api/lotes").json()
        por_id = {l["id"]: l for l in lista}
        assert a["id"] in por_id and b["id"] in por_id
        assert por_id[a["id"]]["competencia"] == "012027"
        assert por_id[b["id"]]["competencia"] == "022027"
        assert "status" in por_id[a["id"]]
        assert "lancamentos" in por_id[a["id"]]

    def test_listar_regras_traz_fornecedor_e_ativo(self, cliente):
        lista = cliente.get("/api/regras").json()
        assert isinstance(lista, list)
        assert len(lista) > 0
        item = lista[0]
        assert "id" in item
        assert "fornecedor" in item
        assert "conta_debito" in item
        assert "ativo" in item
        assert "origem" in item


class TestPortaDeLogin:
    @pytest.fixture(autouse=True)
    def limpa_cookies(self, cliente):
        cliente.cookies.clear()
        yield
        cliente.cookies.clear()

    def test_modo_desligado_nao_pede_conta(self, cliente):
        eu = cliente.get("/api/auth/eu").json()
        assert eu["modo"] == "desligado"
        assert cliente.get("/api/lotes").status_code == 200

    def test_sem_sessao_a_api_fecha(self, cliente, monkeypatch):
        monkeypatch.setenv("AUTH_MODO", "ligado")
        monkeypatch.setenv("AUTH_SECRET", "s" * 48)
        assert cliente.get("/api/lotes").status_code == 401
        assert cliente.get("/api/auth/eu").status_code == 401

    def test_cookie_de_sessao_abre_a_api(self, cliente, monkeypatch):
        monkeypatch.setenv("AUTH_MODO", "ligado")
        monkeypatch.setenv("AUTH_SECRET", "s" * 48)
        from datetime import datetime, timedelta

        from sqlmodel import Session

        from app.autenticacao import COOKIE_SESSAO
        from app.banco import engine
        from app.modelos import SessaoLogin, Usuario

        sid = "sessao-teste-porta"
        with Session(engine) as sessao:
            usuario = Usuario(
                email="contador@projecont.com.br",
                nome="Contador",
                provedor="google",
                provedor_id="g-1",
            )
            sessao.add(usuario)
            sessao.commit()
            sessao.refresh(usuario)
            sessao.add(
                SessaoLogin(
                    id=sid,
                    usuario_id=usuario.id or 0,
                    expira_em=datetime.now() + timedelta(days=1),
                )
            )
            sessao.commit()
        cliente.cookies.set(COOKIE_SESSAO, sid)
        aberta = cliente.get("/api/lotes")
        assert aberta.status_code == 200
        eu = cliente.get("/api/auth/eu").json()
        assert eu["email"] == "contador@projecont.com.br"
        assert eu["modo"] == "ligado"

    def test_origem_estranha_nao_altera_lote(self, cliente, monkeypatch):
        monkeypatch.setenv("AUTH_MODO", "ligado")
        monkeypatch.setenv("AUTH_SECRET", "s" * 48)
        recusada = cliente.post(
            "/api/lotes",
            params={"competencia": "012099"},
            headers={"Origin": "http://evil.example"},
        )
        assert recusada.status_code == 403

    def test_google_redireciona_com_pkce(self, cliente, monkeypatch):
        monkeypatch.setenv("AUTH_MODO", "ligado")
        monkeypatch.setenv("AUTH_SECRET", "s" * 48)
        monkeypatch.setenv("GOOGLE_CLIENT_ID", "gid.apps.googleusercontent.com")
        monkeypatch.setenv("GOOGLE_CLIENT_SECRET", "gsecret")
        monkeypatch.setenv("AUTH_URL_PUBLICA", "http://testserver")
        ida = cliente.get("/api/auth/entrar/google", follow_redirects=False)
        assert ida.status_code == 302
        destino = ida.headers["location"]
        assert destino.startswith("https://accounts.google.com/o/oauth2/v2/auth")
        assert "code_challenge=" in destino
        assert "concilia_oauth" in ida.cookies

    def test_callback_grava_sessao(self, cliente, monkeypatch):
        monkeypatch.setenv("AUTH_MODO", "ligado")
        monkeypatch.setenv("AUTH_SECRET", "s" * 48)
        monkeypatch.setenv("GOOGLE_CLIENT_ID", "gid.apps.googleusercontent.com")
        monkeypatch.setenv("GOOGLE_CLIENT_SECRET", "gsecret")
        monkeypatch.setenv("AUTH_URL_PUBLICA", "http://testserver")
        from unittest.mock import patch
        from urllib.parse import parse_qs, urlparse

        from app.autenticacao import COOKIE_SESSAO, httpx as httpx_auth

        ida = cliente.get("/api/auth/entrar/google", follow_redirects=False)
        estado = parse_qs(urlparse(ida.headers["location"]).query)["state"][0]

        class Token:
            def raise_for_status(self):
                return None

            def json(self):
                return {"access_token": "token-fake"}

        class Perfil:
            def raise_for_status(self):
                return None

            def json(self):
                return {
                    "email": "ana@projecont.com.br",
                    "email_verified": True,
                    "name": "Ana",
                    "sub": "google-sub-1",
                }

        with (
            patch.object(httpx_auth, "post", return_value=Token()),
            patch.object(httpx_auth, "get", return_value=Perfil()),
        ):
            volta = cliente.get(
                "/api/auth/callback/google",
                params={"code": "codigo", "state": estado},
                follow_redirects=False,
            )
        assert volta.status_code == 302
        assert COOKIE_SESSAO in volta.cookies
        aberta = cliente.get("/api/lotes")
        assert aberta.status_code == 200
        assert cliente.get("/api/auth/eu").json()["email"] == "ana@projecont.com.br"

    def test_dominio_fora_da_lista_nao_entra(self, cliente, monkeypatch):
        monkeypatch.setenv("AUTH_MODO", "ligado")
        monkeypatch.setenv("AUTH_SECRET", "s" * 48)
        monkeypatch.setenv("GOOGLE_CLIENT_ID", "gid.apps.googleusercontent.com")
        monkeypatch.setenv("GOOGLE_CLIENT_SECRET", "gsecret")
        monkeypatch.setenv("AUTH_URL_PUBLICA", "http://testserver")
        monkeypatch.setenv("AUTH_DOMINIOS", "projecont.com.br")
        from unittest.mock import patch
        from urllib.parse import parse_qs, urlparse

        from app.autenticacao import httpx as httpx_auth

        ida = cliente.get("/api/auth/entrar/google", follow_redirects=False)
        estado = parse_qs(urlparse(ida.headers["location"]).query)["state"][0]

        class Token:
            def raise_for_status(self):
                return None

            def json(self):
                return {"access_token": "token-fake"}

        class Perfil:
            def raise_for_status(self):
                return None

            def json(self):
                return {
                    "email": "alguem@gmail.com",
                    "email_verified": True,
                    "name": "Alguem",
                    "sub": "google-sub-2",
                }

        with (
            patch.object(httpx_auth, "post", return_value=Token()),
            patch.object(httpx_auth, "get", return_value=Perfil()),
        ):
            volta = cliente.get(
                "/api/auth/callback/google",
                params={"code": "codigo", "state": estado},
                follow_redirects=False,
            )
        assert volta.status_code == 302
        assert "erro=nao_autorizado" in volta.headers["location"]
        assert cliente.get("/api/lotes").status_code == 401

