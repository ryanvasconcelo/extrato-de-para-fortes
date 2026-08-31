"""Gerador do arquivo de importacao FortesERP.

O layout foi derivado empiricamente dos 6 XLSX entregues pelo cliente, porque
specs/RCO010_ImportarLote.pdf nao existe. Ver ADR 0010.

Duas saidas:
  planilha_conferencia  - para o humano revisar, com status e ocorrencias
  arquivo_final         - as 10 colunas, so depois da aprovacao (RF-06.2)
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .modelos import Lancamento, StatusLote

# As 10 colunas, na ordem do arquivo do cliente. H, I e J sao constantes em todas
# as 2.487 linhas historicas e a semantica nao foi confirmada - pergunta aberta ao
# cliente, registrada no handover da Fase 0.
FILIAL = "0001"
CONSTANTE_H = "001"
CONSTANTE_I = "0001"
CONSTANTE_J = "001"
CENTRO_CUSTO_PADRAO = "0001"

FORMATO_DATA = "%d/%m/%Y"

# A linha 1 dos seis arquivos do cliente e hibrida: rotulo nas colunas variaveis,
# constante ja preenchida nas fixas. Nao e cabecalho de planilha nem lancamento -
# e o gabarito do modelo, e sai identica porque o layout do RCO010 nunca chegou
# (ADR 0010). Note o espaco em volta de " Valor ", que existe no original.
LINHA_MODELO = [
    FILIAL,
    "Data",
    "Débito",
    "Crédito",
    " Valor ",
    "Histórico",
    CENTRO_CUSTO_PADRAO,
    CONSTANTE_H,
    CONSTANTE_I,
    CONSTANTE_J,
]


class ExportacaoBloqueada(RuntimeError):
    """Erguida quando se tenta exportar sem aprovacao humana (RF-06.2)."""


def linha_fortes(lancamento: Lancamento, data: date) -> list:
    """Uma linha de lancamento do arquivo final."""
    return [
        lancamento.filial or FILIAL,
        data.strftime(FORMATO_DATA),
        lancamento.conta_debito,
        lancamento.conta_credito,
        round(lancamento.valor, 2),
        lancamento.historico,
        lancamento.centro_custo,
        CONSTANTE_H,
        CONSTANTE_I,
        CONSTANTE_J,
    ]


def arquivo_final(
    status_lote: StatusLote,
    lancamentos: list[tuple[Lancamento, date]],
) -> bytes:
    """Gera o XLSX de importacao. Exige lote APROVADO.

    A trava e aqui, no gerador, e nao apenas na rota: e o unico ponto por onde o
    arquivo pode sair, entao e onde a regra tem que valer.
    """
    if status_lote is not StatusLote.APROVADO:
        raise ExportacaoBloqueada(
            f"Lote em {status_lote.value}. A exportacao exige aprovacao humana "
            f"(status APROVADO)."
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lancamentos"
    ws.append(LINHA_MODELO)
    for lancamento, data in sorted(lancamentos, key=lambda item: item[1]):
        ws.append(linha_fortes(lancamento, data))
    for celula in ws["E"][1:]:
        celula.number_format = "#,##0.00"
    return _bytes(wb)


def planilha_conferencia(
    lancamentos: list[tuple[Lancamento, date]],
    ocorrencias_por_lancamento: dict[int, list[str]],
) -> bytes:
    """Excel de conferencia primeiro, arquivo final so depois de aprovado.

    Padrao de docs/referencia/folha-dealer-checklist-validacao.md.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conferencia"

    cabecalho = [
        "Filial",
        "Data",
        "Conta Debito",
        "Conta Credito",
        "Valor",
        "Historico",
        "Centro Custo",
        "Status",
        "Ocorrencias",
    ]
    ws.append(cabecalho)
    for celula in ws[1]:
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", start_color="FF003048")
        celula.font = Font(bold=True, color="FFFFFFFF")

    destaque = PatternFill("solid", start_color="FFFDEDE4")
    for lancamento, data in sorted(lancamentos, key=lambda item: item[1]):
        codigos = ocorrencias_por_lancamento.get(lancamento.id or -1, [])
        ws.append(
            [
                lancamento.filial or FILIAL,
                data.strftime(FORMATO_DATA),
                lancamento.conta_debito,
                lancamento.conta_credito,
                round(lancamento.valor, 2),
                lancamento.historico,
                lancamento.centro_custo,
                lancamento.status.value,
                ", ".join(codigos),
            ]
        )
        if codigos:
            for celula in ws[ws.max_row]:
                celula.fill = destaque

    for celula in ws["E"][1:]:
        celula.number_format = "#,##0.00"
        celula.alignment = Alignment(horizontal="right")

    larguras = {"A": 8, "B": 12, "C": 20, "D": 20, "E": 14, "F": 58, "G": 12, "H": 12, "I": 36}
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura
    ws.freeze_panes = "A2"

    return _bytes(wb)


def _bytes(wb) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
