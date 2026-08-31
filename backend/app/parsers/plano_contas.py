"""Parser do Plano de Contas e da Base Bancos (mesmo XLSX).

A planilha tem 1.867 linhas, mas apenas 1.750 sao conta: as demais sao cabecalho
de relatorio e linhas em branco. Dessas, 1.516 sao analiticas. Os codigos vem com
digito verificador e a normalizacao e obrigatoria (ADR 0006).
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from ..modelos import ContaBancaria, PlanoContas
from ..normalizacao import conta_e_analitica, normalizar_conta

# As contas correntes ficam sob 1.01.01.02.01.*, com a descricao no formato
# "Banco Itau Ag: 1557 Cc: 98810-0".
_PREFIXO_CONTA_CORRENTE = "1.01.01.02.01."
_DESCRICAO_BANCO = re.compile(
    r"^(?P<banco>.+?)\s+Ag:\s*(?P<agencia>[\w-]+)\s+Cc:\s*(?P<conta>[\w-]+)\s*$",
    re.IGNORECASE,
)


def _primeira_linha_de_dados(ws) -> int:
    """O cabecalho do relatorio ocupa as primeiras linhas; encontra onde comeca."""
    for i, linha in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        primeira = str(linha[0] or "").strip()
        if re.match(r"^\d+(\.\d+)*(-\d)?$", primeira):
            return i
    return 4


def carregar_plano_contas(caminho: Path) -> list[PlanoContas]:
    ws = openpyxl.load_workbook(caminho, data_only=True).active
    inicio = _primeira_linha_de_dados(ws)
    contas: list[PlanoContas] = []
    vistos: set[str] = set()

    for linha in ws.iter_rows(min_row=inicio, values_only=True):
        bruto = str(linha[0] or "").strip()
        if not bruto or not re.match(r"^\d+(\.\d+)*(-\d)?$", bruto):
            continue
        codigo = normalizar_conta(bruto)
        if codigo in vistos:
            continue
        vistos.add(codigo)
        contas.append(
            PlanoContas(
                codigo=codigo,
                codigo_dv=bruto,
                descricao=str(linha[2] or "").strip() if len(linha) > 2 else "",
                natureza=str(linha[3]).strip() if len(linha) > 3 and linha[3] else None,
                reduzido=str(linha[1]).strip() if len(linha) > 1 and linha[1] else None,
                analitica=conta_e_analitica(codigo),
            )
        )
    return contas


def extrair_base_bancos(contas: list[PlanoContas]) -> list[ContaBancaria]:
    """Deriva a Base Bancos do proprio plano de contas.

    Nao ha arquivo separado de Base Bancos; o fluxograma a pressupoe e o plano ja
    tem a informacao na descricao das contas correntes.
    """
    bancos: list[ContaBancaria] = []
    for c in contas:
        if not c.codigo.startswith(_PREFIXO_CONTA_CORRENTE) or not c.analitica:
            continue
        m = _DESCRICAO_BANCO.match(c.descricao)
        if not m:
            continue
        bancos.append(
            ContaBancaria(
                banco=m.group("banco").strip(),
                agencia=m.group("agencia").strip(),
                conta_corrente=m.group("conta").strip(),
                conta_contabil=c.codigo,
            )
        )
    return bancos
